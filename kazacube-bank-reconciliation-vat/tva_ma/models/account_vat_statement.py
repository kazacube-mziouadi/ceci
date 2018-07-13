# -*- encoding: utf-8 -*-

from openerp import models, fields, api, exceptions, _
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
from dateutil import parser
import time
from collections import defaultdict
from itertools import groupby
from operator import itemgetter
import itertools
from cStringIO import StringIO
import json
import base64
import xlwt
import openpyxl
from openpyxl import Workbook
import zipfile
from openpyxl.styles import PatternFill, Alignment, Font, Color, colors

class account_vat_payment_line(models.Model):
    _name = "account.vat.payment.line"


    vat_statement_id = fields.Many2one(comodel_name="account.vat.statement", string="Déclaration TVA", ondelete="cascade")
    invoice_id = fields.Many2one(comodel_name="account.invoice", string="Facture")
    invoice_number = fields.Char(string="Numéro facture", required=False, )
    invoice_date = fields.Date(string="Date facture", required=False, )
    payment_date = fields.Date(string="Date règlement", required=False, )
    reglement_method_id = fields.Many2one(comodel_name="account.invoice.tva.reglement", string="Méthode", required=False, )
    bank_statement_date = fields.Date(string="Date du relevé", required=False, )
    bank_statement_line_date = fields.Date(string="Date de l'opération", required=False, )
    bank_statement_line_id = fields.Many2one(comodel_name="account.bank.statement.line", string="Ligne du relevé bancaire", required=False, )
    invoice_state = fields.Char(string="Statut facture", required=False, )
    invoice_currency_id = fields.Many2one(comodel_name="res.currency", string="Devise")
    invoice_partner_id = fields.Many2one(comodel_name="res.partner", string="Partenaire")
    partner_name = fields.Char(string="Nom", required=False, )
    partner_ifu = fields.Char(string="IF", required=False, )
    partner_ice = fields.Char(string="ICE", required=False, )
    invoice_partner_ifu = fields.Char(string="IF", required=False, )
    invoice_amount_untaxed = fields.Float(string="Facture HT",  required=False, )
    invoice_amount_tax = fields.Float(string="Facture Taxe",  required=False, )
    payment_amount_tax = fields.Float(string="Encaissé Taxe",  required=False, )
    payment_amount_untaxed = fields.Float(string="Encaissé HT",  required=False, )
    invoice_amount_total = fields.Float(string="facture TTC",  required=False, )
    payment_amount = fields.Float(string="Encaissé TTC",  required=False, )
    tax_id = fields.Many2one(comodel_name="account.tax", string="Taxe")
    taxes = fields.Text(string="Taxes", required=False, )
    invoice_move_line_id = fields.Many2one(comodel_name="account.move.line", string="Ecriture comptable tiers associée à la facture", required=False, )
    payment_move_line_id = fields.Many2one(comodel_name="account.move.line", string="Ecriture comptable tiers associée au règlement", required=False, )
    state = fields.Selection(string="Statut", selection=[('draft', 'Brouillon'), ('confirmed', 'Validée'), ('closed', 'Clôturée') ], default='draft')
    vat_type = fields.Selection(string="Type TVA", selection=[('sale', 'Imposable'), ('purchase', 'Déductible') ])
    full_partial = fields.Selection(string="Total/Partiel", selection=[('full', 'Total'), ('partial', 'Partiel') ])
    nature = fields.Char(string="Nature", required=False, )
    payment_method_name = fields.Char(string="Méthode règlement", compute="_get_payment_method_name", store=True )
    currency_name = fields.Char(string="Devise", compute="_get_currency", store=True )
    vat_name = fields.Char(string="Devise", compute="_get_vat_name", store=True )

    @api.one
    @api.depends('tax_id')
    def _get_vat_name(self):
        if self.tax_id:
            self.vat_name = self.tax_id.name
        else:
            self.vat_name = False

    @api.one
    @api.depends('reglement_method_id')
    def _get_payment_method_name(self):
        if self.reglement_method_id:
            self.payment_method_name = self.reglement_method_id.name
        else:
            self.payment_method_name = False

    @api.one
    @api.depends('invoice_currency_id')
    def _get_currency(self):
        if self.invoice_currency_id:
            self.currency_name = self.invoice_currency_id.name
        else:
            self.currency_name = False



    @api.onchange('tax_id','amount')
    def onchange_tax_id(self):
        if self.tax_id and self.payment_amount:
            self.payment_amount_untaxed = self.payment_amount /(1+( self.tax_id.amount))
            self.payment_amount_tax = self.payment_amount - self.payment_amount_untaxed


account_vat_payment_line()



class account_vat_bank_line(models.Model):
    _name = "account.vat.bank.line"

    name = fields.Char(string="Libellé", required=False, )
    ref = fields.Char(string="Référence", required=False, )
    vat_statement_id = fields.Many2one(comodel_name="account.vat.statement", string="Déclaration TVA", ondelete="cascade")
    invoice_number = fields.Char(string="Numéro facture", required=False, )
    invoice_currency_id = fields.Many2one(comodel_name="res.currency", string="Devise")
    invoice_date = fields.Date(string="Date facture", required=False, )
    payment_date = fields.Date(string="Date règlement", required=False, )
    reglement_method_id = fields.Many2one(comodel_name="account.invoice.tva.reglement", string="Méthode", required=False, )
    date = fields.Date(string="Date opération", required=False, )
    invoice_partner_id = fields.Many2one(comodel_name="res.partner", string="Partenaire")
    partner_name = fields.Char(string="Nom", required=False, )
    partner_ifu = fields.Char(string="IF", required=False, )
    partner_ice = fields.Char(string="ICE", required=False, )
    invoice_amount_untaxed = fields.Float(string="Facture HT",  required=False, )
    invoice_amount_tax = fields.Float(string="Facture Taxe",  required=False, )
    invoice_amount_total = fields.Float(string="facture TTC",  required=False, )
    amount_untaxed = fields.Float(string="Montant HT",  required=False, )
    amount_tax = fields.Float(string="Montant taxe",  required=False, )
    amount = fields.Float(string="Montant TTC",  required=False, )
    tax_id = fields.Many2one(comodel_name="account.tax", string="Taxe", required=False, )
    bank_statement_line_id = fields.Many2one(comodel_name="account.bank.statement.line", string="Ligne du relevé bancaire", required=False, )
    state = fields.Selection(string="Statut", selection=[('draft', 'Brouillon'), ('confirmed', 'Validée'), ('closed', 'Clôturée') ], default='draft')
    vat_type = fields.Selection(string="Type TVA", selection=[('sale', 'Imposable'), ('purchase', 'Déductible') ])
    nature = fields.Char(string="Nature", required=False, )
    payment_method_name = fields.Char(string="Méthode règlement", compute="_get_payment_method_name", store=True )
    currency_name = fields.Char(string="Devise", compute="_get_currency", store=True )
    vat_name = fields.Char(string="Devise", compute="_get_vat_name", store=True )

    @api.one
    @api.depends('tax_id')
    def _get_vat_name(self):
        if self.tax_id:
            self.vat_name = self.tax_id.name
        else:
            self.vat_name = False

    @api.one
    @api.depends('reglement_method_id')
    def _get_payment_method_name(self):
        if self.reglement_method_id:
            self.payment_method_name = self.reglement_method_id.name
        else:
            self.payment_method_name = False

    @api.one
    @api.depends('invoice_currency_id')
    def _get_currency(self):
        if self.invoice_currency_id:
            self.currency_name = self.invoice_currency_id.name
        else:
            self.currency_name = False

    @api.onchange('tax_id','amount')
    def onchange_tax_id(self):
        if self.tax_id and self.amount:
            self.amount_untaxed = self.amount /(1+( self.tax_id.amount))
            self.amount_tax = self.amount - self.amount_untaxed

    @api.onchange('invoice_partner_id')
    def onchange_invoice_partner_id(self):
        if self.invoice_partner_id:
            self.partner_ifu = self.invoice_partner_id.ifu
            self.partner_ice = self.invoice_partner_id.ice
            self.partner_name = self.invoice_partner_id.name

account_vat_bank_line()

class account_vat_move_line(models.Model):
    _name = "account.vat.move.line"

    name = fields.Char(string="Libellé", required=False, )
    ref = fields.Char(string="Référence", required=False, )
    vat_statement_id = fields.Many2one(comodel_name="account.vat.statement", string="Déclaration TVA", ondelete="cascade")
    invoice_number = fields.Char(string="Numéro facture", required=False, )
    invoice_currency_id = fields.Many2one(comodel_name="res.currency", string="Devise")
    invoice_date = fields.Date(string="Date facture", required=False, )
    payment_date = fields.Date(string="Date règlement", required=False, )
    reglement_method_id = fields.Many2one(comodel_name="account.invoice.tva.reglement", string="Méthode", required=False, )
    date = fields.Date(string="Date opération", required=False, )
    invoice_partner_id = fields.Many2one(comodel_name="res.partner", string="Partenaire")
    partner_name = fields.Char(string="Nom", required=False, )
    partner_ifu = fields.Char(string="IF", required=False, )
    partner_ice = fields.Char(string="ICE", required=False, )
    invoice_amount_untaxed = fields.Float(string="Facture HT",  required=False, )
    invoice_amount_tax = fields.Float(string="Facture Taxe",  required=False, )
    invoice_amount_total = fields.Float(string="facture TTC",  required=False, )
    amount_untaxed = fields.Float(string="Montant HT",  required=False, )
    amount_tax = fields.Float(string="Montant taxe",  required=False, )
    amount = fields.Float(string="Montant TTC",  required=False, )
    payment_move_line_id = fields.Many2one(comodel_name="account.move.line", string="Ecriture comptable tiers associée au règlement", required=False, )
    bank_move_line_id = fields.Many2one(comodel_name="account.move.line", string="Ecriture comptable de banque associée au règlement", required=False, )
    bank_statement_line_id = fields.Many2one(comodel_name="account.bank.statement.line", string="Ligne du relevé bancaire", required=False, )
    tax_id = fields.Many2one(comodel_name="account.tax", string="Taxe", required=False, )
    state = fields.Selection(string="Statut", selection=[('draft', 'Brouillon'), ('confirmed', 'Validée'), ('closed', 'Clôturée') ], default='draft')
    vat_type = fields.Selection(string="Type TVA", selection=[('sale', 'Imposable'), ('purchase', 'Déductible') ])
    move_type = fields.Selection(string="Type", selection=[('isolated_reconciled', 'Lettrées sans facture'), ('isolated', 'Isolées') ])
    nature = fields.Char(string="Nature", required=False, )
    payment_method_name = fields.Char(string="Méthode règlement", compute="_get_payment_method_name", store=True )
    currency_name = fields.Char(string="Devise", compute="_get_currency", store=True )
    vat_name = fields.Char(string="Devise", compute="_get_vat_name", store=True )

    @api.one
    @api.depends('tax_id')
    def _get_vat_name(self):
        if self.tax_id:
            self.vat_name = self.tax_id.name
        else:
            self.vat_name = False

    @api.one
    @api.depends('reglement_method_id')
    def _get_payment_method_name(self):
        if self.reglement_method_id:
            self.payment_method_name = self.reglement_method_id.name
        else:
            self.payment_method_name = False

    @api.one
    @api.depends('invoice_currency_id')
    def _get_currency(self):
        if self.invoice_currency_id:
            self.currency_name = self.invoice_currency_id.name
        else:
            self.currency_name = False

    @api.onchange('tax_id','amount')
    def onchange_tax_id(self):
        if self.tax_id and self.amount:
            self.amount_untaxed = self.amount /(1+( self.tax_id.amount))
            self.amount_tax = self.amount - self.amount_untaxed

    @api.onchange('invoice_partner_id')
    def onchange_invoice_partner_id(self):
        if self.invoice_partner_id:
            self.partner_ifu = self.invoice_partner_id.ifu
            self.partner_ice = self.invoice_partner_id.ice
            self.partner_name = self.invoice_partner_id.name

account_vat_move_line()

class account_vat_statement(models.Model):
    _name = "account.vat.statement"

    name = fields.Char(string="Nom", required=True)
    fiscalyear_id = fields.Many2one(comodel_name="account.fiscalyear", string="Année fiscale", required=True, )
    period_id = fields.Many2one(comodel_name="account.period.tva", string="Période", required=True, )
    date_debut = fields.Date( compute="onchange_period_id",string="Date début", required=True)
    date_fin = fields.Date(compute="onchange_period_id",string="Date fin", required=True)
    tva_sale = fields.Boolean(string="CA Imposable")
    tva_purchase = fields.Boolean(string="TVA Deductible")
    bank_reconcile = fields.Boolean(string="Récupérer les écritures dont les relevés bancaires sont rapprochés seulement",help="Si coché, la TVA n'est calculée que pour les encaissements qui sont rapprochés avec des relevés bancaires")
    type_period_tva = fields.Selection(string="Périodicité", selection=[('1', 'Mensuelle'), ('3', 'Trimestrielle')], required=True)
    state = fields.Selection(string="Statut", selection=[('draft', 'Brouillon'), ('confirmed', 'Validée'), ('closed', 'Clôturée') ], default='draft')
    vat_ids1 = fields.One2many(comodel_name="account.vat.payment.line", inverse_name="vat_statement_id", string="TVA sur encaissements", required=False, )
    vat_ids2 = fields.One2many(comodel_name="account.vat.bank.line", inverse_name="vat_statement_id", string="TVA sur relevés bancaires", required=False, )
    vat_ids3 = fields.One2many(comodel_name="account.vat.move.line", inverse_name="vat_statement_id", string="TVA sur écritures comptables", required=False, )
    nb_payment_lines = fields.Integer(string="TVAs (payments)", required=False,compute="_get_nb_payment_lines" )
    nb_full_payment_lines = fields.Integer(string="TVAs (Full payments)", required=False,compute="_get_nb_full_payment_lines" )
    nb_partial_payment_lines = fields.Integer(string="TVAs (Partial payments)", required=False,compute="_get_nb_partial_payment_lines" )
    nb_bank_lines = fields.Integer(string="TVAs (bank statements)", required=False,compute="_get_nb_bank_lines" )
    nb_move_lines = fields.Integer(string="TVAs (move lines)", required=False,compute="_get_nb_move_lines" )
    nb_reconciled_move_lines = fields.Integer(string="TVAs (reconciled move lines)", required=False,compute="_get_nb_reconciled_move_lines" )
    nb_isolated_move_lines = fields.Integer(string="TVAs (isolated move lines)", required=False,compute="_get_nb_isolated_move_lines" )
    mp_indicator=fields.Integer(string='entrer ID du mode de paiement ')
    choose_mp=fields.Selection([('espece', 'Espèce'), ('cheque', 'Cheque'), ('virement', 'Virement'),('prelevement','Prélèvement'),('effet','Effet'),('compensation','Compensation'),('autres','Autres')])
    total_ca_imposable = fields.Float(string="TOTAL CA IMPOSABLE TTC", compute="_get_total_ca_imposable", store=False )
    total_ca_imposable_untaxed = fields.Float(string="TOTAL CA IMPOSABLE HT", compute="_get_total_ca_imposable_untaxed", store=False )

    total_ca_non_imposable = fields.Float(string="TOTAL CA NON IMPOSABLE", compute="_get_total_ca_non_imposable", store=False )


    total_vat = fields.Float(string="TVA FACTUREE", compute="_get_total_vat", store=False )

    @api.one
    @api.depends('vat_ids1','vat_ids1.payment_amount_tax','vat_ids2','vat_ids2.amount_tax','vat_ids3','vat_ids3.amount_tax')
    def _get_total_vat(self):
        total = 0
        for line in self.vat_ids1:
            total+=line.payment_amount_tax
        for line in self.vat_ids2:
            total+=line.amount_tax
        for line in self.vat_ids3:
            total+=line.amount_tax
        self.total_vat = total

    @api.one
    @api.depends('vat_ids1','vat_ids1.payment_amount','vat_ids2','vat_ids2.amount','vat_ids3','vat_ids3.amount')
    def _get_total_ca_imposable(self):
        total = 0
        for line in self.vat_ids1:
            if line.tax_id:
                total+=line.payment_amount
        for line in self.vat_ids2:
            if line.tax_id:
                total+=line.amount
        for line in self.vat_ids3:
            if line.tax_id:
                total+=line.amount
        self.total_ca_imposable = total

    @api.one
    @api.depends('vat_ids1','vat_ids1.payment_amount_untaxed','vat_ids2','vat_ids2.amount_untaxed','vat_ids3','vat_ids3.amount_untaxed')
    def _get_total_ca_non_imposable(self):
        total = 0
        for line in self.vat_ids1:
            if not line.tax_id:
                total+=line.payment_amount_untaxed
        for line in self.vat_ids2:
            if not line.tax_id:
                total+=line.amount_untaxed
        for line in self.vat_ids3:
            if not line.tax_id:
                total+=line.amount_untaxed
        self.total_ca_non_imposable = total

    @api.one
    @api.depends('vat_ids1','vat_ids1.payment_amount_untaxed','vat_ids2','vat_ids2.amount_untaxed','vat_ids3','vat_ids3.amount_untaxed')
    def _get_total_ca_imposable_untaxed(self):
        total = 0
        for line in self.vat_ids1:
            if line.tax_id:
                total+=line.payment_amount_untaxed
        for line in self.vat_ids2:
            if line.tax_id:
                total+=line.amount_untaxed
        for line in self.vat_ids3:
            if line.tax_id:
                total+=line.amount_untaxed
        self.total_ca_imposable_untaxed = total


    @api.one
    def _get_nb_payment_lines(self):
        lines = self.env['account.vat.payment.line'].search([('vat_statement_id','=',self.id)])
        self.nb_payment_lines = lines and len(lines) or 0

    @api.one
    def _get_nb_full_payment_lines(self):
        lines = self.env['account.vat.payment.line'].search([('vat_statement_id','=',self.id),('full_partial','=','full')])
        self.nb_full_payment_lines = lines and len(lines) or 0

    @api.one
    def _get_nb_partial_payment_lines(self):
        lines = self.env['account.vat.payment.line'].search([('vat_statement_id','=',self.id),('full_partial','=','partial')])
        self.nb_partial_payment_lines = lines and len(lines) or 0

    @api.one
    def _get_nb_bank_lines(self):
        lines = self.env['account.vat.bank.line'].search([('vat_statement_id','=',self.id)])
        self.nb_bank_lines = lines and len(lines) or 0

    @api.one
    def _get_nb_move_lines(self):
        lines = self.env['account.vat.move.line'].search([('vat_statement_id','=',self.id)])
        self.nb_move_lines = lines and len(lines) or 0

    @api.one
    def _get_nb_reconciled_move_lines(self):
        lines = self.env['account.vat.move.line'].search([('vat_statement_id','=',self.id),('move_type','=','isolated_reconciled')])
        self.nb_reconciled_move_lines = lines and len(lines) or 0

    @api.one
    def _get_nb_isolated_move_lines(self):
        lines = self.env['account.vat.move.line'].search([('vat_statement_id','=',self.id),('move_type','=','isolated')])
        self.nb_isolated_move_lines = lines and len(lines) or 0

    @api.one
    @api.constrains('period_id')
    def _check_unique_statement_per_period(self):
        vat_statements = False
        filter= False
        if self.tva_sale:
            filter = "tva_sale"
        else:
            filter = "tva_purchase"
        vat_statements = self.env['account.vat.statement'].search([('state','in',['confirmed','closed']),('period_id','=',self.period_id.id),(filter,'=',True)])
        if vat_statements:
            raise exceptions.ValidationError(_("Cette période est dèja declarée"))



    @api.one
    def gen_vat(self):
        self.generate_vat_from_payments()
        self.generate_vat_from_bank_statements()
        return True

    @api.multi
    def check_bank_reconciliation(self, payment_move_id):

        print "move id ====",payment_move_id

        req = "SELECT aml.id AS bank_move_line_id " \
              " FROM account_move_line AS aml, account_move AS am, account_journal AS aj " \
              " WHERE aml.move_id = am.id AND aj.id = aml.journal_id" \
              " AND aml.move_id = "+str(payment_move_id)+" " \
              " AND aj.type = 'bank' " \
              " AND aml.move_stat_id IS NOT NULL "

        self._cr.execute(req)
        print "req check bank reconcile =",req
        dict_result = self._cr.dictfetchall()
        if len(dict_result) >= 1:
            req = "SELECT absl.id AS statement_line_id, absl.date AS statement_line_date, " \
                    " abs.id AS statement_id, abs.date AS statement_date  " \
                    " FROM account_bank_statement_line AS absl, account_bank_statement AS abs " \
                    " WHERE absl.statement_id = abs.id " \
                    " AND absl.move_stat_id IS NOT NULL  " \
                    " AND abs.date <= '"+self.date_fin+"' " \
                    " AND absl.id IN  " \
                    " (SELECT bank_statement_line_id FROM move_statement_relation WHERE account_move_line_id = "+str(dict_result[0]['bank_move_line_id'])+" ) "

            self._cr.execute(req)
            result = self._cr.dictfetchall()
            print "REEEEEEEEEESULT ===",result
            if len(result) >= 1:
                return {'status':'OK', 'statement_date':result[0]['statement_date'], 'statement_line_date':result[0]['statement_line_date'], 'statement_line_id':result[0]['statement_line_id']}
            else:
                return {'status':'NOK'}
        else:
            return {'status':'NOK'}

    @api.one
    def _create_lines_for_invoices_with_single_tax(self,dict_elem, column_amount):

        print "CREATING SINGLE"

        tax_id = False
        tax_percentage = False
        req="SELECT ailt.tax_id AS tax_id, act.amount AS tax_percentage " \
            " FROM account_invoice_line_tax AS ailt, account_tax AS act  " \
            " WHERE ailt.tax_id = act.id " \
            " AND ailt.invoice_line_id IN (SELECT ail.id FROM account_invoice_line AS ail WHERE ail.invoice_id = "+str(dict_elem['invoice_id'])+")"

        self._cr.execute(req)
        result = self._cr.dictfetchall()
        if len(result) >=1:
            tax_id = result[0] and result[0]['tax_id'] or False
            tax_percentage = result[0] and result[0]['tax_percentage'] or False

        if dict_elem['reconcile_id']:
            reconcile_column = 'reconcile_id'
        else:
            reconcile_column = 'reconcile_partial_id'

        req = "SELECT aml.id AS payment_move_line_id, aml.move_id AS payment_move_id, aml.date AS payment_date, aml.rapprocher AS rapprocher, " \
                  " aml."+column_amount+" AS payment_amount, aitr.id AS payment_method_id " \
                  " FROM account_move_line AS aml, account_invoice_tva_reglement AS aitr " \
                  " WHERE aml.reglement_method_id = aitr.id  " \
                  " AND aml.date <= '"+self.date_fin+"' " \
                  " AND aml."+reconcile_column+" = '"+str(dict_elem['reconcile_id'] or dict_elem['reconcile_partial_id'])+"' " \
                  " AND aml.id <> "+str(dict_elem['invoice_move_line_id'])
        print req

        self._cr.execute(req)
        dict_result2 = self._cr.dictfetchall()

        for payment_move_line in dict_result2:

            reconciliation_status = False
            if self.bank_reconcile:
                reconciliation_status = self.check_bank_reconciliation(payment_move_line['payment_move_id'])
                print "reconciliation_status ==",reconciliation_status
                if reconciliation_status['status']  == 'NOK':
                    continue


            vals={
                'vat_statement_id':self.id,
                'invoice_move_line_id':dict_elem['invoice_move_line_id'],
                'payment_move_line_id':payment_move_line['payment_move_line_id'],
                'full_partial':dict_elem['reconcile_id'] and 'full' or 'partial',
                'invoice_date':dict_elem['invoice_date'],
                'payment_date':payment_move_line['payment_date'],
                'bank_statement_date':reconciliation_status and reconciliation_status['statement_date'] or False,
                'bank_statement_line_date':reconciliation_status and reconciliation_status['statement_line_date'] or False,
                'bank_statement_line_id':reconciliation_status and reconciliation_status['statement_line_id'] or False,
                'invoice_id':dict_elem['invoice_id'],
                'invoice_number':dict_elem['invoice_number'],
                'invoice_state':dict_elem['invoice_state'],
                'invoice_partner_id':dict_elem['partner_id'],
                'partner_name':dict_elem['partner_name'],
                'partner_ifu':dict_elem['partner_ifu'],
                'partner_ice':dict_elem['partner_ice'],
                'invoice_currency_id':dict_elem['currency_id'],
                'invoice_amount_tax':dict_elem['invoice_amount_tax'],
                'invoice_amount_total':dict_elem['invoice_amount_total'],
                'invoice_amount_untaxed':dict_elem['invoice_amount_untaxed'],
                'payment_amount':payment_move_line['payment_amount'],
                'tax_id':tax_id,
                'payment_amount_untaxed':tax_percentage and (payment_move_line['payment_amount']/(1+tax_percentage)) or 0.0,
                'vat_type':self.tva_sale and 'sale' or 'purchase',
                'reglement_method_id':payment_move_line['payment_method_id'],
            }
            vals['payment_amount_tax'] = vals['payment_amount_untaxed']*tax_percentage

            self.env['account.vat.payment.line'].create(vals)

        return True

    @api.one
    def _create_lines_for_invoices_with_multiple_taxes(self,dict_elem, column_amount):

        print "CREATING MULTIPLE"

        if dict_elem['reconcile_id']:
            reconcile_column = 'reconcile_id'
        else:
            reconcile_column = 'reconcile_partial_id'

        req = "SELECT aml.id AS payment_move_line_id, aml.move_id AS payment_move_id, aml.date AS payment_date, aml.rapprocher AS rapprocher, " \
                  " aml."+column_amount+" AS payment_amount, aitr.id AS payment_method_id " \
                  " FROM account_move_line AS aml, account_invoice_tva_reglement AS aitr " \
                  " WHERE aml.reglement_method_id = aitr.id  " \
                  " AND aml.date <= '"+self.date_fin+"' " \
                  " AND aml."+reconcile_column+" = '"+str(dict_elem['reconcile_id'] or dict_elem['reconcile_partial_id'])+"' " \
                  " AND aml.id <> "+str(dict_elem['invoice_move_line_id'])
        print req
        self._cr.execute(req)
        dict_result2 = self._cr.dictfetchall()

        for payment_move_line in dict_result2:

            reconciliation_status = False
            if self.bank_reconcile:
                reconciliation_status = self.check_bank_reconciliation(payment_move_line['payment_move_id'])
                print "reconciliation_status =",reconciliation_status
                if reconciliation_status['status'] == 'NOK':
                    continue

            req="SELECT ait.tax_id AS tax_id , act.amount AS tax_percentage, ait.amount AS tax_amount, ait.base AS amount_untaxed, ait.base+ait.amount AS amount_total " \
                " FROM account_invoice_tax as ait, account_tax AS act " \
                " WHERE ait.tax_id = act.id " \
                " AND ait.invoice_id = "+str(dict_elem['invoice_id'])+" "
            self._cr.execute(req)
            print "req =",req
            invoice_taxes_dict = self._cr.dictfetchall()

            for tax_line in invoice_taxes_dict:

                ratio = payment_move_line['payment_amount']/dict_elem['invoice_amount_total']
                prorata_ttc = ratio * tax_line['amount_total']

                vals={
                    'vat_statement_id':self.id,
                    'invoice_move_line_id':dict_elem['invoice_move_line_id'],
                    'payment_move_line_id':payment_move_line['payment_move_line_id'],
                    'full_partial':dict_elem['reconcile_id'] and 'full' or 'partial',
                    'invoice_date':dict_elem['invoice_date'],
                    'payment_date':payment_move_line['payment_date'],
                    'bank_statement_date':reconciliation_status and reconciliation_status['statement_date'] or False,
                    'bank_statement_line_date':reconciliation_status and reconciliation_status['statement_line_date'] or False,
                    'bank_statement_line_id':reconciliation_status and reconciliation_status['statement_line_id'] or False,
                    'invoice_id':dict_elem['invoice_id'],
                    'invoice_number':dict_elem['invoice_number'],
                    'invoice_state':dict_elem['invoice_state'],
                    'invoice_partner_id':dict_elem['partner_id'],
                    'partner_name':dict_elem['partner_name'],
                    'partner_ifu':dict_elem['partner_ifu'],
                    'partner_ice':dict_elem['partner_ice'],
                    'invoice_currency_id':dict_elem['currency_id'],
                    'invoice_amount_tax':dict_elem['invoice_amount_tax'],
                    'invoice_amount_total':dict_elem['invoice_amount_total'],
                    'invoice_amount_untaxed':dict_elem['invoice_amount_untaxed'],
                    'payment_amount':prorata_ttc,
                    'payment_amount_tax': prorata_ttc-(prorata_ttc/(1+ tax_line['tax_percentage'])),
                    'tax_id':tax_line['tax_id'],
                    'vat_type':self.tva_sale and 'sale' or 'purchase',
                    'reglement_method_id':payment_move_line['payment_method_id'],
                }

                vals['payment_amount_untaxed'] = vals['payment_amount'] - vals['payment_amount_tax']

                self.env['account.vat.payment.line'].create(vals)

        return True


    @api.one
    def generate_vat_from_payments(self):

        #self._cr.execute("DELETE FROM account_vat_payment_line WHERE vat_statement_id = '"+str(self.id)+"'")
        #MZ######
        ###ice_number
        ###num_ifu
        req = "SELECT ai.date_invoice AS invoice_date, aml.date AS payment_date,ai.id AS invoice_id, ai.state AS invoice_state, ai.partner_id AS partner_id," \
              "ai.currency_id AS currency_id, ai.amount_total AS invoice_amount_total, ai.amount_untaxed AS invoice_amount_untaxed, " \
              "ai.amount_tax AS invoice_amount_tax, rp.name AS partner_name, rp.ice_number AS partner_ice, rp.num_if AS partner_ifu, " \
              "aml.id AS invoice_move_line_id, aml.reconcile_id AS reconcile_id, aml.reconcile_partial_id AS reconcile_partial_id, " \
              "aml.rapprocher AS rapprocher, ai.number AS invoice_number " \
              " FROM account_invoice AS ai, account_move AS am, account_move_line AS aml, res_partner AS rp  " \
              " WHERE ai.id = am.invoice_id AND am.id = aml.move_id AND ai.partner_id = rp.id " \
              " AND ai.account_id = aml.account_id " \
              " AND (aml.reconcile_id IS NOT NULL OR aml.reconcile_partial_id IS NOT NULL) " \
              " AND ai.state <> 'draft' " \
              " AND aml.date <= '"+self.date_fin+"' " \
              " AND aml.id NOT IN (SELECT invoice_move_line_id FROM account_vat_payment_line WHERE state <> 'draft') "

        column_amount = False
        if self.tva_sale:
            req+=" and ai.type = 'out_invoice' "
            column_amount = 'credit'
        else:
            req+=" and ai.type = 'in_invoice' "
            column_amount = 'debit'

        print req
        self._cr.execute(req)
        dict_result = self._cr.dictfetchall()
        print dict_result

        for elem in dict_result:
            req = "SELECT ait.id AS tax_line_id FROM account_invoice_tax AS ait WHERE ait.invoice_id = '"+str(elem['invoice_id'])+"'"
            self._cr.execute(req)
            result = self._cr.dictfetchall()
            if len(result) == 1:
                self._create_lines_for_invoices_with_single_tax(elem, column_amount)
            else:
                self._create_lines_for_invoices_with_multiple_taxes(elem, column_amount)

        return True






    @api.one
    def generate_vat_from_bank_statements(self):

        #self._cr.execute("DELETE FROM account_vat_bank_line WHERE vat_statement_id = '"+str(self.id)+"'")

        req = " SELECT bsl.name AS name, bsl.id AS id, bsl.date AS date, bsl.amount AS amount " \
              " FROM account_bank_statement_line AS bsl, account_bank_statement AS bs, account_journal AS aj " \
              " WHERE bsl.statement_id = bs.id AND bs.journal_id = aj.id" \
              " AND bs.state <> 'draft'  " \
              " AND aj.type = 'bank'  " \
              " AND bsl.date <= '"+str(self.date_fin)+"' " \
              " AND bsl.move_stat_id IS NULL " \
              " AND bsl.id NOT IN (SELECT bank_statement_line_id FROM account_vat_bank_line WHERE state <> 'draft') "

        if self.tva_sale:
            req+=" and bsl.amount > 0"
            taxes = self.env['account.tax'].search([('name','=','TVA 20% VENTES')])
        else:
            req+=" and bsl.amount < 0"
            taxes = self.env['account.tax'].search([('name','=','TVA 20% ACHATS')])
        self._cr.execute(req)
        dict_result = self._cr.dictfetchall()

        default_tax_id = False
        if taxes:
            default_tax_id = taxes[0].id

        for elem in dict_result:
            vals = {
                'name':elem['name'],
                'vat_statement_id':self.id,
                'bank_statement_line_id':elem['id'],
                'date':elem['date'],
                'amount':elem['amount'],
                'tax_id':default_tax_id,
                'amount_untaxed':elem['amount']/(1+taxes[0].amount),
                'vat_type':self.tva_sale and 'sale' or 'purchase',
            }
            vals['amount_tax']= vals['amount']-vals['amount_untaxed']
            self.env['account.vat.bank.line'].create(vals)

        return True

    @api.one
    def _get_isolated_reconciled_move_lines(self, wizard_id):

        req = "SELECT aml.move_id as move_id, aml.id AS payment_move_line_id " \
              " FROM account_move_line AS aml, account_move AS am, account_journal AS aj " \
              " WHERE aml.move_id = am.id AND aml.journal_id = aj.id " \
              " AND am.state = 'posted' " \
              " AND (aj.type = 'bank' OR aj.type = 'cash') " \
              " AND (aml.reconcile_id IS NOT NULL OR aml.reconcile_partial_id IS NOT NULL) " \
              " AND aml.date <= '"+str(self.date_fin)+"' " \
              " AND " \
              " (" \
              " (aml.reconcile_id IN (SELECT aml.reconcile_id " \
                                    " FROM account_move_line AS aml, account_move AS am, account_journal AS aj " \
                                    " WHERE aml.journal_id = aj.id AND aml.move_id = am.id " \
                                    " AND am.state = 'posted' " \
                                    " AND aml.reconcile_id IS NOT NULL" \
                                    " AND aml.date <= '"+str(self.date_fin)+"' " \
                                    " AND (aj.type <> 'cash') " \
                                    " AND (aj.type <> 'bank') " \
                                    " AND am.invoice_id IS NULL " \
                                    ")) " \
              " OR (aml.reconcile_partial_id IN (SELECT aml.reconcile_partial_id " \
                                    " FROM account_move_line AS aml, account_move AS am, account_journal AS aj " \
                                    " WHERE aml.journal_id = aj.id AND aml.move_id = am.id " \
                                    " AND am.state = 'posted' " \
                                    " AND aml.reconcile_partial_id IS NOT NULL" \
                                    " AND aml.date <= '"+str(self.date_fin)+"' " \
                                    " AND (aj.type <> 'cash') " \
                                    " AND (aj.type <> 'bank') " \
                                    " AND am.invoice_id IS NULL " \
                                    ")) " \
              ") " \
              " AND aml.id NOT IN (SELECT payment_move_line_id FROM account_vat_move_line WHERE state <> 'draft') "


        if self.tva_sale:
            req+=" AND aml.credit > 0 "
        else:
            req+=" AND aml.debit > 0 "

        print "TTTTTTTTTTTTTTTTTTTTTTT ISOLATED RECONCILED"
        print req
        self._cr.execute(req)
        dict_result = self._cr.dictfetchall()
        print dict_result

        for payment_move_line in dict_result:

            req= "SELECT aml.name AS name, aml.ref AS ref, aml.journal_id AS journal_id, aj.type AS journal_type,  " \
                 " aml.account_id AS account_id, aml.partner_id AS partner_id, " \
                 " aml.credit AS credit, aml.debit AS debit, aml.date AS move_line_date, " \
                 " aml.id AS bank_move_line_id, aml.rapprocher AS rapprocher, aml.move_stat_id AS move_stat_id  " \
                 " FROM account_move_line AS aml, account_move AS am, account_journal AS aj " \
                 " WHERE aml.move_id = am.id AND aml.journal_id = aj.id " \
                 " AND am.id = "+str(payment_move_line['move_id'])+" " \
                 " AND aml.id <> "+str(payment_move_line['payment_move_line_id'])+" "

            self._cr.execute(req)
            result = self._cr.dictfetchall()

            for bank_move_line in result:
                print "isolated reconciled move_stat_id ",bank_move_line['move_stat_id']
                print bank_move_line['bank_move_line_id']
                if self.bank_reconcile and bank_move_line['journal_type'] == 'bank' and bank_move_line['move_stat_id'] is None:
                    print "isolated reconciled continue"
                    continue
                move_statement_relations = self.env['move.statement.relation'].search([('id','=',bank_move_line['move_stat_id']),('account_move_line_id','=',bank_move_line['bank_move_line_id'])])
                vals = {
                    'wizard_id':wizard_id,
                    'name':bank_move_line['name'],
                    'ref':bank_move_line['ref'],
                    'date':bank_move_line['move_line_date'],
                    'journal_id':bank_move_line['journal_id'],
                    'account_id':bank_move_line['account_id'],
                    'partner_id':bank_move_line['partner_id'],
                    'bank_move_line_id':bank_move_line['bank_move_line_id'],
                    'payment_move_line_id':payment_move_line['payment_move_line_id'],
                    'bank_statement_line_id':move_statement_relations and move_statement_relations[0].bank_statement_line_id.id or False,
                    'debit':bank_move_line['debit'],
                    'credit':bank_move_line['credit'],
                    'move_type':'isolated_reconciled',
                }
                self.env['move.lines.vat.line'].create(vals)

        return True

    @api.one
    def _get_isolated_move_lines(self, wizard_id):

        req = " SELECT aml.name AS name, aml.ref AS ref, aml.journal_id AS journal_id, aj.type AS journal_type,  " \
              " aml.account_id AS account_id, aml.partner_id AS partner_id, " \
              " aml.credit AS credit, aml.debit AS debit, aml.date AS move_line_date, " \
              " aml.id AS bank_move_line_id, aml.rapprocher AS rapprocher, aml.move_stat_id AS move_stat_id  " \
              " FROM account_move_line AS aml, account_move AS am, account_journal AS aj, account_account AS acc " \
              " WHERE aml.move_id = am.id AND aml.journal_id = aj.id AND aml.account_id = acc.id " \
              " AND am.state = 'posted' " \
              " AND (aj.type = 'bank' OR aj.type = 'cash') " \
              " AND aml.date <= '"+str(self.date_fin)+"' " \
              " AND acc.type = 'liquidity' " \
              " AND aml.move_id IN " \
              "( " \
              " SELECT DISTINCT am.id AS move_id " \
              " FROM account_move_line AS aml, account_move AS am, account_account AS acc, account_journal AS aj " \
              " WHERE aml.move_id = am.id AND aml.account_id = acc.id AND aml.journal_id = aj.id " \
              " AND am.state = 'posted' " \
              " AND (aj.type = 'bank' OR aj.type = 'cash') " \
              " AND aml.date <= '"+str(self.date_fin)+"' " \
              " AND acc.type <> 'receivable' " \
              " AND acc.type <> 'payable' " \
              " AND acc.type <> 'liquidity' " \
              " ) " \
              " AND aml.id NOT IN (SELECT bank_move_line_id FROM account_vat_move_line WHERE state <> 'draft') "

        if self.tva_sale:
            req+=" AND aml.debit > 0 "
        else:
            req+=" AND aml.credit > 0 "

        print "LLLLLLLLLLLLLLLLLLLLLL ISOLATED"
        print req
        self._cr.execute(req)
        dict_result = self._cr.dictfetchall()
        print dict_result

        for bank_move_line in dict_result:
                print "isolated move_stat_id ",bank_move_line['move_stat_id']
                print bank_move_line['bank_move_line_id']
                if self.bank_reconcile and bank_move_line['journal_type'] == 'bank' and bank_move_line['move_stat_id'] is None:
                    print "isolated continue"
                    continue
                move_statement_relations = self.env['move.statement.relation'].search([('id','=',bank_move_line['move_stat_id']),('account_move_line_id','=',bank_move_line['bank_move_line_id'])])
                vals = {
                    'wizard_id':wizard_id,
                    'name':bank_move_line['name'],
                    'ref':bank_move_line['ref'],
                    'date':bank_move_line['move_line_date'],
                    'journal_id':bank_move_line['journal_id'],
                    'account_id':bank_move_line['account_id'],
                    'partner_id':bank_move_line['partner_id'],
                    'bank_move_line_id':bank_move_line['bank_move_line_id'],
                    'payment_move_line_id':False,
                    'bank_statement_line_id':move_statement_relations and move_statement_relations[0].bank_statement_line_id.id or False,
                    'debit':bank_move_line['debit'],
                    'credit':bank_move_line['credit'],
                    'move_type':'isolated',
                }
                self.env['move.lines.vat.line'].create(vals)

        return True


    @api.multi
    def action_pick_moves(self):
        wizard_id = self.env['move.lines.vat.wizard'].create({'vat_statement_id':self.id})

        self._get_isolated_reconciled_move_lines(wizard_id.id)
        self._get_isolated_move_lines(wizard_id.id)

        return {
            'name':_("Séléction des lignes à déclarer"),
            'view_mode': 'form',
            'view_id': False,
            'view_type': 'form',
            'res_model': 'move.lines.vat.wizard',
            'res_id':wizard_id.id,
            'type': 'ir.actions.act_window',
            'nodestroy': True,
            'target': 'new',
            'domain': '[]',
            }


    @api.onchange('period_id','tva_sale','tva_purchase')
    def write_statement_name(self):
        if self.period_id:
            if self.tva_sale:
                self.name = self.period_id.title1
            if self.tva_purchase:
                self.name = self.period_id.title2

    @api.onchange('tva_sale')
    def onchange_tva_sale(self):
        if self.tva_sale and self.tva_purchase:
            self.tva_purchase = False

    @api.onchange('tva_purchase')
    def onchange_tva_purchase(self):
        if self.tva_purchase and self.tva_sale:
            self.tva_sale = False

    @api.onchange('type_period_tva')
    def onchange_type_period_tva(self):
        res = {'domain': {}}
        if self.fiscalyear_id and self.type_period_tva:
            periods = self.env['account.period.tva'].search([('fiscalyear_id','=',self.fiscalyear_id.id),('type','=',self.type_period_tva)])
            if periods:
                res['domain']['period_id'] = [('id','in',[p.id for p in periods])]
            else:
                res['domain']['period_id'] = [('id','in',[])]
            return res
        else:
            self.period_id = False
            res['domain']['period_id'] = [('id','in',[])]
            return res

    @api.one
    @api.depends('period_id')
    def onchange_period_id(self):
        if self.period_id:
            self.date_debut = self.period_id.date_start
            self.date_fin = self.period_id.date_stop
        else:
            self.date_debut = False
            self.date_fin = False


    @api.one
    def set_closed(self):
        for v in self.vat_ids1:
            v.state = 'closed'
        for v in self.vat_ids2:
            v.state = 'closed'
        for v in self.vat_ids3:
            v.state = 'closed'
        self.state = 'closed'

    @api.one
    def set_draft(self):
        for v in self.vat_ids1:
            v.state = 'draft'
        for v in self.vat_ids2:
            v.state = 'draft'
        for v in self.vat_ids3:
            v.state = 'draft'
        self.state = 'draft'

    @api.one
    def set_confirmed(self):
        for v in self.vat_ids1:
            v.state = 'confirmed'
        for v in self.vat_ids2:
            v.state = 'confirmed'
        for v in self.vat_ids3:
            v.state = 'confirmed'
        self.state = 'confirmed'


    @api.multi
    def _check_if_bank_statement_line_is_vat_posted(self,statement_line_id):
        req="SELECT bank_statement_line_id FROM account_vat_payment_line WHERE state <> 'draft' AND bank_statement_line_id = "+str(statement_line_id)+" "
        req+=" UNION ALL "
        req+="SELECT bank_statement_line_id FROM account_vat_bank_line WHERE state <> 'draft' AND bank_statement_line_id = "+str(statement_line_id)+" "
        req+=" UNION ALL "
        req+="SELECT bank_statement_line_id FROM account_vat_move_line WHERE state <> 'draft' AND bank_statement_line_id = "+str(statement_line_id)+" "
        self._cr.execute(req)
        print req
        result = self._cr.fetchall()
        print result
        if not result:
            return False
        else:
            return True

    @api.multi
    def button_report_xml(self):
        results = {}

        periode= self.period_id.code
        bperiode = periode[1:]
        periodicite=int(bperiode)


        #################### Requête encaissements/décaissements, relevés bancaires, écritures comptables manuelles #######################
        reqs = "SELECT invoice_date, invoice_number, partner_name, partner_ifu, partner_ice, " \
               " nature, currency_name, invoice_amount_untaxed, vat_name, invoice_amount_tax, invoice_amount_total, " \
               " payment_amount_untaxed, payment_amount_tax, payment_amount, payment_date, payment_method_name, account_invoice_tva_reglement.id_xml    " \
               " FROM account_vat_payment_line INNER JOIN account_invoice_tva_reglement ON account_vat_payment_line.reglement_method_id=account_invoice_tva_reglement.id " \
               " WHERE vat_statement_id = " + str(self.id) + " "

        reqs += " UNION ALL "
        reqs += "SELECT invoice_date, invoice_number, partner_name, partner_ifu, partner_ice, " \
                " nature, currency_name, invoice_amount_untaxed, vat_name, invoice_amount_tax, invoice_amount_total, " \
                " amount_untaxed, amount_tax, amount, date, payment_method_name ,account_invoice_tva_reglement.id_xml    " \
                " FROM account_vat_bank_line INNER JOIN account_invoice_tva_reglement ON account_vat_bank_line.reglement_method_id=account_invoice_tva_reglement.id " \
                " WHERE vat_statement_id = " + str(self.id) + "  "
        reqs += " UNION ALL "
        reqs += "SELECT invoice_date, invoice_number, partner_name, partner_ifu, partner_ice, " \
                " nature, currency_name, invoice_amount_untaxed, vat_name, invoice_amount_tax, invoice_amount_total, " \
                " amount_untaxed, amount_tax, amount, date, payment_method_name  ,account_invoice_tva_reglement.id_xml   " \
                " FROM account_vat_move_line INNER JOIN account_invoice_tva_reglement ON account_vat_move_line.reglement_method_id=account_invoice_tva_reglement.id " \
                " WHERE vat_statement_id = " + str(self.id) + "  " \
                                                              " Order by vat_name "

        self._cr.execute(reqs)
        print reqs
        resultats = self._cr.fetchall()

        qweb_context_tva = {

            'date_year': datetime.now().strftime('%Y'),
            'type_period':1 if self.type_period_tva == '1' else 2,
            'identifiantFiscal': self.fiscalyear_id.company_id.vat,
            'lines': resultats,
            'periodicity':periodicite,
        }
        # ***************************************#
        results['TVA'] = self.generate_zip_file('tva_xml_template', qweb_context_tva, 'tva_')
        res = self.env["report.tva.xml.wizard"].create(
            {'TVA_file': results["TVA"]["data"], 'tva_filename': results["TVA"]["filename"],})

        return {
            'name': "fichier XML de déclaration TVA",
            'view_mode': 'form',
            'view_id': False,
            'view_type': 'form',
            'res_model': 'report.tva.xml.wizard',
            'res_id': res.id,
            'type': 'ir.actions.act_window',
            'nodestroy': True,
            'target': 'new',


        }



    def generate_zip_file(self, xml_template, qweb_context, filename_prefix):
        result = {}

        template = self.env['ir.ui.view'].search([('name', '=', xml_template)])
        xml_text = self.pool['ir.ui.view'].render(self._cr, self._uid, template.xml_id, qweb_context, engine='ir.qweb',
                                                  context=None)
        xml_text = '<?xml version="1.0" encoding="UTF-8"?>\n' + xml_text

        file = open(
            "/tmp/" + filename_prefix + self.fiscalyear_id.name + "_"  + ".xml",
            "wb")
        file.write(xml_text)
        file.close()

        zp = zipfile.ZipFile("/tmp/" + filename_prefix + self.fiscalyear_id.name + "_" + ".zip", "w")

        zp.write("/tmp/" + filename_prefix + self.fiscalyear_id.name + "_"  + ".xml",
                 arcname=filename_prefix + self.fiscalyear_id.name + "_" + ".xml")
        zp.close()

        file = open(
            "/tmp/" + filename_prefix + self.fiscalyear_id.name + "_" +  ".zip",
            "rb")
        out = file.read().encode('base64', 'strict')
        file.close()

        result = {'data': out,
                  'filename': filename_prefix + self.fiscalyear_id.name + "_" +  ".zip"}

        return result









    @api.multi
    def button_report_excel(self):
        date_day = datetime.now().strftime('%d-%m-%Y')
        i=10
        j=1
        buf=StringIO()
        #############formatage des cellules###############################"
        ft= Font(name='Calibri',size=15,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
        ft_header= Font(name='Calibri',size=15,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='7191DC')
        ###############################definition de fichier excel##############
        wb = Workbook()
        ws = wb.active
        ####################################Remplir la feuille excel avec les données########################################
        date_day = datetime.now().strftime('%d-%m-%Y')
        txt1 = unicode('Déclaration TVA', "utf8")
        title1 = txt1+' '+str(self.name)
        ws.cell(column=4, row=1).value=title1
        ws.cell(column=4, row=1).font=ft



        ce=ws.cell(column=1, row=4, value=unicode('Société', "utf8"))
        ce=ws.cell(column=2, row=4, value=self.env.user.company_id.name)

        ce=ws.cell(column=4, row=4, value=unicode('Adresse', "utf8"))
        ce=ws.cell(column=5, row=4, value=self.env.user.partner_id.street or " ")
        ce=ws.cell(column=6, row=4, value=+self.env.user.partner_id.city or " ")


        #################### Header ############################
        header=['Date facture','Numéro facture','Nom partenaire','IF','ICE','Nature opération','Devise Facture','Montant HT facture','Taux de la TVA','Montant TVA facture','Montant TTC facture',
                'Montant HT encaissement', 'Montant TVA encaissement', 'Montant TTC encaissement','Date règlement','Méthode règlement']
        h=1
        for elem in header:
            ce=ws.cell(column=h, row=9, value=unicode(elem, "utf8"))
            h+=1
        #################### Requête encaissements/décaissements, relevés bancaires, écritures comptables manuelles #######################
        req = "SELECT invoice_date, invoice_number, partner_name, partner_ifu, partner_ice, " \
              " nature, currency_name, invoice_amount_untaxed, vat_name, invoice_amount_tax, invoice_amount_total, " \
              " payment_amount_untaxed, payment_amount_tax, payment_amount, payment_date, payment_method_name    " \
              " FROM account_vat_payment_line " \
              " WHERE vat_statement_id = "+str(self.id)+" "
        req+=" UNION ALL "
        req += "SELECT invoice_date, invoice_number, partner_name, partner_ifu, partner_ice, " \
              " nature, currency_name, invoice_amount_untaxed, vat_name, invoice_amount_tax, invoice_amount_total, " \
              " amount_untaxed, amount_tax, amount, date, payment_method_name    " \
              " FROM account_vat_bank_line " \
              " WHERE vat_statement_id = "+str(self.id)+"  "
        req+=" UNION ALL "
        req += "SELECT invoice_date, invoice_number, partner_name, partner_ifu, partner_ice, " \
              " nature, currency_name, invoice_amount_untaxed, vat_name, invoice_amount_tax, invoice_amount_total, " \
              " amount_untaxed, amount_tax, amount, date, payment_method_name    " \
              " FROM account_vat_move_line " \
              " WHERE vat_statement_id = "+str(self.id)+"  " \
              " Order by vat_name "

        self._cr.execute(req)
        print req
        result = self._cr.fetchall()
        for elements in result:
            for elem in elements:
                ce=ws.cell(column=j, row=i, value=elem)
                j+=1
            i+=1
            j=1

        #TOTAL CA IMPOSABLE/DEDUCTIBLE
        ce=ws.cell(column=4, row=i+3, value=unicode(self.tva_sale and 'TOTAL CA IMPOSABLE' or 'TOTAL CA DEDUCTIBLE', "utf8"))
        ce=ws.cell(column=12, row=i+6, value=self.total_ca_imposable_untaxed)
        ce=ws.cell(column=13, row=i+6, value=self.total_vat)
        ce=ws.cell(column=14, row=i+6, value=self.total_ca_imposable)
        # entête tableau
        ce=ws.cell(column=3, row=i+5, value=unicode('TTC', "utf8"))
        ce=ws.cell(column=4, row=i+5, value=unicode('HT', "utf8"))
        ### TOTAL CA IMPOSABLE/DEDUCTIBLE
        ce=ws.cell(column=1, row=i+6, value=unicode(self.tva_sale and 'TOTAL CA IMPOSABLE' or 'TOTAL CA DEDUCTIBLE', "utf8"))
        ce=ws.cell(column=3, row=i+6, value=self.total_ca_imposable)
        ce=ws.cell(column=4, row=i+6, value=self.total_ca_imposable_untaxed)
        #### TOTAL CA NON IMPOSABLE/DEDUCTIBLE
        ce=ws.cell(column=1, row=i+7, value=unicode(self.tva_sale and 'TOTAL CA NON IMPOSABLE' or 'TOTAL CA NON DEDUCTIBLE', "utf8"))
        ce=ws.cell(column=3, row=i+7, value=self.total_ca_non_imposable)
        ce=ws.cell(column=4, row=i+7, value=self.total_ca_non_imposable)
        ##### TVA FACTUREE #####
        ce=ws.cell(column=1, row=i+8, value=unicode('TVA FACTUREE', "utf8"))
        ce=ws.cell(column=3, row=i+8, value=self.total_vat)

        ###################################Enregistrement du fichier####################################
        wb.save(buf)

        fichier = self.name+"_"+str(date_day)+".xlsx"
        out=base64.encodestring(buf.getvalue())
        buf.close()
        vals={'data':out,'name_file':fichier}
        wizard_id = self.pool.get("report.excel.wizard").create(self._cr, self._uid, vals, context=self._context)
        return {
            'name':_("Rapport Excel de déclaration TVA"),
            'view_mode': 'form',
            'view_id': False,
            'view_type': 'form',
            'res_model': 'report.excel.wizard',
            'res_id':wizard_id,
            'type': 'ir.actions.act_window',
            'target': 'new',
            'domain': '[]',
            }


account_vat_statement()

class report_excel_wizard(models.TransientModel):
    _name = 'report.excel.wizard'

    name_file = fields.Char(string="Nom Fichier")
    data = fields.Binary(string="Fichier")

report_excel_wizard()

